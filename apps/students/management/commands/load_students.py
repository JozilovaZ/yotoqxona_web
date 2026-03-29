import openpyxl
from django.core.management.base import BaseCommand
from buildings.models import Building, Floor, Room
from students.models import Student


class Command(BaseCommand):
    help = "TATU SF 2-son DTM Excel faylidan talaba ma'lumotlarini yuklash"

    def handle(self, *args, **options):
        # Building yaratish
        building, created = Building.objects.get_or_create(
            name="TATU SF 2-son DTM",
            defaults={
                'address': "Samarqand shahri",
                'monthly_price': 450000,
                'is_active': True,
            }
        )
        if created:
            self.stdout.write(self.style.SUCCESS(f"Bino yaratildi: {building.name}"))
        else:
            self.stdout.write(f"Bino mavjud: {building.name}")

        # Qavatlarni yaratish
        floors_data = {
            1: 'female',  # 1-qavat ayollar
            2: 'male',    # 2-qavat erkaklar
            3: 'male',    # 3-qavat erkaklar
            4: 'male',    # 4-qavat erkaklar
        }
        floors = {}
        for num, gender in floors_data.items():
            floor, created = Floor.objects.get_or_create(
                building=building,
                number=num,
                defaults={'gender': gender, 'is_active': True}
            )
            floors[num] = floor
            if created:
                self.stdout.write(f"  Qavat yaratildi: {num}-qavat ({gender})")

        # Excel faylni o'qish
        wb = openpyxl.load_workbook('jadval_1/TATU SF 2-son dxsh .xlsx')
        ws = wb[wb.sheetnames[0]]

        students_created = 0
        students_updated = 0
        rooms_created = 0
        last_room_number = None

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            vals = {cell.column: cell.value for cell in row if cell.value is not None}

            # Ism bo'lmasa, o'tkazib yuborish
            full_name_raw = vals.get(2)
            if not full_name_raw or not isinstance(full_name_raw, str):
                continue

            full_name_raw = full_name_raw.strip()
            if not full_name_raw:
                continue

            # Ma'lumotlarni olish
            faculty = vals.get(3, 'TTKT')
            course = vals.get(4, 1)
            floor_num = vals.get(5)
            room_number = vals.get(6)
            passport = vals.get(7, '')
            jshir = vals.get(8, '')
            phone = vals.get(9, '')
            summa = vals.get(10, 450000)

            # Xona raqami - agar bo'sh bo'lsa, oxirgi xonani ishlatish
            if room_number is not None:
                last_room_number = str(int(room_number)) if isinstance(room_number, (int, float)) else str(room_number)
            current_room_number = last_room_number

            if not current_room_number or not floor_num:
                continue

            floor_num = int(floor_num)

            # Ismni bo'laklarga ajratish
            name_parts = full_name_raw.split()
            last_name = name_parts[0] if len(name_parts) > 0 else ''
            first_name = name_parts[1] if len(name_parts) > 1 else ''
            middle_name = ' '.join(name_parts[2:]) if len(name_parts) > 2 else ''

            # Jinsni aniqlash
            gender = 'female' if floor_num == 1 else 'male'

            # Telefon raqamni formatlash
            if phone:
                phone = f"+998{str(int(phone))}" if isinstance(phone, (int, float)) else str(phone)

            # Passport raqam - student_id sifatida
            if passport:
                passport = str(passport).strip()
            else:
                passport = f"NO_ID_{vals.get(1, 0)}"

            # JSHIR
            if jshir:
                jshir = str(int(jshir)) if isinstance(jshir, (int, float)) else str(jshir)

            # Floor va Room olish/yaratish
            floor = floors.get(floor_num)
            if not floor:
                continue

            room, r_created = Room.objects.get_or_create(
                floor=floor,
                number=current_room_number,
                defaults={
                    'capacity': 4,
                    'room_type': 'standard',
                    'status': 'available',
                    'is_active': True,
                }
            )
            if r_created:
                rooms_created += 1

            # Talabani yaratish yoki yangilash
            student, s_created = Student.objects.update_or_create(
                student_id=passport,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'middle_name': middle_name,
                    'gender': gender,
                    'phone': phone,
                    'faculty': faculty if faculty else 'TTKT',
                    'course': int(course) if course else 1,
                    'room': room,
                    'is_active': True,
                    'notes': f"JSHIR: {jshir}" if jshir else '',
                }
            )

            if s_created:
                students_created += 1
            else:
                students_updated += 1

        # Xonalar statusini yangilash
        for room in Room.objects.filter(floor__building=building):
            room.update_status()

        self.stdout.write(self.style.SUCCESS(
            f"\nNatija:\n"
            f"  Xonalar yaratildi: {rooms_created}\n"
            f"  Talabalar yaratildi: {students_created}\n"
            f"  Talabalar yangilandi: {students_updated}\n"
            f"  Jami talabalar: {Student.objects.filter(room__floor__building=building).count()}"
        ))
