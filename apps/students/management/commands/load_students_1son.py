import openpyxl
from django.core.management.base import BaseCommand
from buildings.models import Building, Floor, Room
from students.models import Student


class Command(BaseCommand):
    help = "TATU SF 1-son Excel faylidan talaba ma'lumotlarini yuklash"

    def handle(self, *args, **options):
        # Building yaratish
        building, created = Building.objects.get_or_create(
            name="TATU SF 1-son DTM",
            defaults={
                'address': "Samarqand shahri",
                'monthly_price': 450000,
                'is_active': True,
            }
        )
        if created:
            self.stdout.write(self.style.SUCCESS(f"Bino yaratildi: {building.name}"))
        else:
            self.stdout.write(f"Bino mavjud: {building.name}")

        # Qavatlarni yaratish
        floors_data = {
            1: 'female',  # 1-qavat (asosan ayollar)
            2: 'male',    # 2-qavat erkaklar
            3: 'male',    # 3-qavat erkaklar
            4: 'male',    # 4-qavat erkaklar
        }
        floors = {}
        for num, gender in floors_data.items():
            floor, created = Floor.objects.get_or_create(
                building=building,
                number=num,
                defaults={'gender': gender, 'is_active': True}
            )
            floors[num] = floor
            if created:
                self.stdout.write(f"  Qavat yaratildi: {num}-qavat ({gender})")

        # Excel faylni o'qish
        wb = openpyxl.load_workbook('jadval_1/TATU SF 1-son .xlsx')
        ws = wb[wb.sheetnames[0]]

        students_created = 0
        students_updated = 0
        rooms_created = 0

        # Ayol ismlarining oxirlari (Uzbek)
        female_suffixes = ('ova', 'yeva', 'eva', 'qizi', 'ovna', 'xon')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            vals = {cell.column: cell.value for cell in row if cell.value is not None}

            # Ism bo'lmasa, o'tkazib yuborish
            full_name_raw = vals.get(2)
            if not full_name_raw or not isinstance(full_name_raw, str):
                continue

            full_name_raw = full_name_raw.strip()
            if not full_name_raw:
                continue

            # Ma'lumotlarni olish
            index_num = vals.get(1, 0)
            address = vals.get(3, '')
            group = vals.get(4, '')
            phone = vals.get(5, '')
            room_raw = vals.get(6, '')
            debt_info = vals.get(7, '')
            payment_info = vals.get(8, '')

            # Xona raqamini ajratish: "101-Xona" -> "101"
            if not room_raw or not isinstance(room_raw, str):
                continue
            room_number = room_raw.lower().replace('-xona', '').replace('-xpna', '').strip()
            try:
                room_num_int = int(room_number)
            except ValueError:
                self.stdout.write(self.style.WARNING(f"  Noto'g'ri xona: {room_raw} (qator {index_num})"))
                continue

            # Qavatni xona raqamidan aniqlash (1xx=1-qavat, 2xx=2-qavat, ...)
            floor_num = room_num_int // 100
            if floor_num not in floors:
                self.stdout.write(self.style.WARNING(
                    f"  Noto'g'ri qavat: {floor_num} (xona {room_raw}, qator {index_num})"
                ))
                continue

            # Ismni bo'laklarga ajratish
            name_parts = full_name_raw.split()
            last_name = name_parts[0] if len(name_parts) > 0 else ''
            first_name = name_parts[1] if len(name_parts) > 1 else ''
            middle_name = ' '.join(name_parts[2:]) if len(name_parts) > 2 else ''

            # Jinsni aniqlash - familiya oxiriga qarab
            gender = 'male'
            name_lower = last_name.lower()
            for suffix in female_suffixes:
                if name_lower.endswith(suffix):
                    gender = 'female'
                    break

            # Telefon raqamni formatlash
            if phone:
                phone = f"+998{str(int(phone))}" if isinstance(phone, (int, float)) else str(phone)

            # Student ID yaratish (passport yo'q, shuning uchun index asosida)
            student_id = f"1SON_{int(index_num):03d}"

            # Kurs raqamini guruh nomidan aniqlash
            course = 1
            if group and isinstance(group, str):
                # Guruh formati: MT22-09, KI25-04, DI23-15 etc.
                # 2 raqam - yil, masalan 22 = 2022, course = 2026 - 2022 = 4
                import re
                match = re.search(r'(\d{2})-', group)
                if match:
                    year = int(match.group(1))
                    # 2026 yil uchun kursni hisoblash
                    course = 26 - year
                    if course < 1:
                        course = 1
                    elif course > 6:
                        course = 6

            # Fakultetni guruh nomidan aniqlash
            faculty = 'TTKT'
            if group and isinstance(group, str):
                prefix = re.split(r'\d', group.strip())[0].strip()
                if prefix:
                    faculty = prefix

            # Qarz ma'lumotlarini notes ga saqlash
            notes_parts = []
            if address:
                notes_parts.append(f"Yashash joyi: {address}")
            if debt_info and debt_info != '-':
                notes_parts.append(f"Qarz: {debt_info}")
            if payment_info and payment_info != '-':
                notes_parts.append(f"To'lov: {payment_info}")
            notes = ' | '.join(notes_parts)

            # Floor va Room olish/yaratish
            floor = floors[floor_num]

            room, r_created = Room.objects.get_or_create(
                floor=floor,
                number=str(room_num_int),
                defaults={
                    'capacity': 4,
                    'room_type': 'standard',
                    'status': 'available',
                    'is_active': True,
                }
            )
            if r_created:
                rooms_created += 1

            # Talabani yaratish yoki yangilash
            student, s_created = Student.objects.update_or_create(
                student_id=student_id,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'middle_name': middle_name,
                    'gender': gender,
                    'phone': phone,
                    'faculty': faculty,
                    'course': course,
                    'group': group.strip() if group else '',
                    'room': room,
                    'is_active': True,
                    'notes': notes,
                }
            )

            if s_created:
                students_created += 1
            else:
                students_updated += 1

        # Xonalar statusini yangilash
        for room in Room.objects.filter(floor__building=building):
            room.update_status()

        self.stdout.write(self.style.SUCCESS(
            f"\nNatija:\n"
            f"  Xonalar yaratildi: {rooms_created}\n"
            f"  Talabalar yaratildi: {students_created}\n"
            f"  Talabalar yangilandi: {students_updated}\n"
            f"  Jami talabalar: {Student.objects.filter(room__floor__building=building).count()}"
        ))
